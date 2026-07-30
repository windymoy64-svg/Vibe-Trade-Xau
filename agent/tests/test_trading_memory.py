"""Unit tests for Trading Memory Schema v1.0 and its writers."""

from __future__ import annotations

import json
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from src.trading.memory import (
    ExcelMemoryWriter, JsonMemoryWriter, MemoryEngine, TradingMemory, WORKSHEETS,
)


def _memory(**identity_updates: object) -> TradingMemory:
    identity = {
        "trade_id": "trade-001", "replay_id": "replay-001", "experiment_id": "exp-001",
        "walkforward_id": "wf-001", "strategy_version": "1.0.0",
        "runtime_config_version": "4", "git_commit": "abc123",
        "timestamp": datetime(2026, 7, 30, 12, tzinfo=timezone.utc), "environment": "test",
        **identity_updates,
    }
    return MemoryEngine().create({
        "identity": identity,
        "market_context": {"symbol": "XAUUSD", "timeframe": "1h", "spread": 0.2, "atr": 12.5},
        "technical_snapshot": {
            "ema": {"ema_20": 2310.5}, "rsi": 57.0, "macd": {"line": 1.2}, "atr": 12.5,
            "features": {"future_indicator": 42},
        },
        "decision_snapshot": {
            "direction": "BUY", "confidence": 0.8, "entry_reasons": ["EMA_TREND"],
            "signal_components": {"trend": 0.9}, "expected_probability": 0.7,
            "expected_rr": 2.0, "expected_holding_time": 3600,
        },
        "risk_snapshot": {
            "risk_percent": 1, "lot_size": 0.1, "entry": 2300, "stop_loss": 2290,
            "take_profit": 2320, "rr": 2, "atr_stop": 10, "position_size": 0.1,
        },
        "execution_snapshot": {
            "executed_entry": 2300.2, "executed_exit": 2320, "slippage": 0.2,
            "spread": 0.2, "commission": 1.5, "swap": 0, "execution_latency": 15, "duration": 3600,
        },
        "result_snapshot": {
            "outcome": "WIN", "exit_reason": "TP", "profit_usd": 198,
            "profit_percent": 1.98, "profit_r": 1.98, "pips": 198,
        },
        "post_mortem": {"observations": ["Hit Resistance"]},
        "validation_snapshot": {"replay_version": "1.0"},
    })


def test_memory_creation_defaults_and_extensible_features() -> None:
    memory = _memory()
    assert memory.schema_version == "1.0"
    assert memory.technical_snapshot.features["future_indicator"] == 42
    assert memory.lesson.model_dump() == {
        "lesson": None, "lesson_type": None, "confidence": None,
        "created_by": None, "status": None,
    }


def test_unique_memory_ids() -> None:
    assert _memory().identity.memory_id != _memory().identity.memory_id


def test_json_serialization_and_append(tmp_path) -> None:
    path = tmp_path / "journal.json"
    writer = JsonMemoryWriter(path)
    first, second = _memory(), _memory(trade_id="trade-002")
    writer.write(first)
    writer.write(second)
    payload = json.loads(path.read_text(encoding="utf-8"))
    assert payload["schema_version"] == "1.0"
    assert len(payload["memories"]) == 2
    assert TradingMemory.model_validate(payload["memories"][0]) == first
    with pytest.raises(ValueError, match="already exists"):
        writer.write(first)


def test_excel_export(tmp_path) -> None:
    path = ExcelMemoryWriter(tmp_path / "Trading_Memory.xlsx").write(_memory())
    workbook = load_workbook(path, read_only=True)
    assert tuple(workbook.sheetnames) == WORKSHEETS
    assert workbook["Dashboard"]["B3"].value == 1
    assert workbook["Trade History"].max_row == 2
    assert workbook["Technical Snapshot"]["A2"].value is not None


def test_schema_validation_and_immutability() -> None:
    with pytest.raises(ValidationError, match="timezone-aware"):
        _memory(timestamp=datetime(2026, 7, 30, 12))
    invalid = _memory().model_dump(mode="json")
    invalid["decision_snapshot"]["direction"] = "HOLD"
    with pytest.raises(ValidationError):
        TradingMemory.model_validate(invalid)
    memory = _memory()
    with pytest.raises(ValidationError):
        memory.identity.trade_id = "changed"  # type: ignore[misc]
