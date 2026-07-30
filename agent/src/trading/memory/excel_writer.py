"""Human-readable normalized Trading Memory workbook writer."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .memory_schema import TradingMemory
from .memory_writer import MemoryWriter

WORKSHEETS = (
    "Dashboard", "Trade History", "Market Context", "Technical Snapshot",
    "Fundamental Snapshot", "Decision Snapshot", "Risk Snapshot",
    "Execution Snapshot", "Results", "Lessons", "Validation",
)


def _cell(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


class ExcelMemoryWriter(MemoryWriter):
    def __init__(self, path: str | Path = "Trading_Memory.xlsx") -> None:
        self.path = Path(path)

    @staticmethod
    def _append(sheet: Worksheet, memory_id: str, values: dict[str, Any]) -> None:
        headers = ["memory_id", *values.keys()]
        if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
            for column, header in enumerate(headers, start=1):
                sheet.cell(1, column, header)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
        sheet.append([memory_id, *(_cell(value) for value in values.values())])

    def _workbook(self) -> Workbook:
        if self.path.exists():
            return load_workbook(self.path)
        workbook = Workbook()
        workbook.active.title = WORKSHEETS[0]
        for title in WORKSHEETS[1:]:
            workbook.create_sheet(title)
        dashboard = workbook["Dashboard"]
        dashboard.append(["Trading Memory Dashboard", "Value"])
        dashboard.append(["Schema Version", "1.0"])
        dashboard.append(["Total Memories", 0])
        for cell in dashboard[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        return workbook

    def write(self, memory: TradingMemory) -> Path:
        workbook = self._workbook()
        memory_id = str(memory.identity.memory_id)
        history = workbook["Trade History"]
        if history.max_row > 1 and any(str(row[0].value) == memory_id for row in history.iter_rows(min_row=2)):
            raise ValueError(f"memory_id already exists: {memory_id}")

        identity = memory.identity.model_dump(mode="json")
        identity.pop("memory_id")
        identity["schema_version"] = memory.schema_version
        identity["post_mortem"] = list(memory.post_mortem.observations)
        self._append(history, memory_id, identity)
        sections = (
            ("Market Context", memory.market_context),
            ("Technical Snapshot", memory.technical_snapshot),
            ("Fundamental Snapshot", memory.fundamental_snapshot),
            ("Decision Snapshot", memory.decision_snapshot),
            ("Risk Snapshot", memory.risk_snapshot),
            ("Execution Snapshot", memory.execution_snapshot),
            ("Results", memory.result_snapshot),
            ("Lessons", memory.lesson),
            ("Validation", memory.validation_snapshot),
        )
        for sheet_name, snapshot in sections:
            self._append(workbook[sheet_name], memory_id, snapshot.model_dump(mode="json"))
        workbook["Dashboard"]["B3"] = workbook["Trade History"].max_row - 1
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.path)
        return self.path
